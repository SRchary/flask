from openpyxl.styles import Font


def apply_header_format(workbook, worksheet, dataframe, row_num, column_num):
    merge_format = workbook.add_format({
        'bold': 1, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'fg_color': '#efefef'
    })
    for col_num, value in enumerate(dataframe.columns.values):
        worksheet.write(row_num, column_num + col_num, value, merge_format)


def apply_border(workbook, worksheet, range_value):
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    cell_format.set_text_wrap()
    worksheet.conditional_format(range_value, {'type': 'no_blanks', 'format': cell_format})
    worksheet.conditional_format(range_value, {'type': 'blanks', 'format': cell_format})
    return worksheet


def merge_cells(workbook, worksheet, cell_range, title):
    merge_format = workbook.add_format({
        'bold': 1, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'fg_color': '#efefef'
    })
    worksheet.merge_range(cell_range, title, merge_format)
    return worksheet


def title_format_cells(workbook, worksheet, cell_range, title):
    title_format = workbook.add_format({
        'align': 'center', 'valign': 'vcenter'
    })
    title_format.set_font_size(16)
    worksheet.merge_range(cell_range, title, title_format)
    return worksheet


def format_oh_conductor_sheet(workbook, worksheet, df):
    """
    This function will format oh_conductor Sheet in FERC excel book.
    :param workbook: Workbook Object
    :param worksheet: Oh Conductor Sheet Object for applying sheet level styles
    :param df: Data Frame object
    :return: None

    """
    merge_format = workbook.add_format({
        'bold': 1, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'fg_color': '#efefef'
    })
    index = df.shape[0] + 1
    cell_format = workbook.add_format()
    cell_format.set_align('center')
    cell_format.set_align('vcenter')
    #cell_format.set_font_name('Calibri')
    worksheet.set_column('A:N', 30, cell_format)
    apply_border(workbook, worksheet, "A1:M{}".format(index))
    total_length = df['CIRCUIT_MI'].sum()
    worksheet.write("H{}".format(index + 2), "Grand Total", merge_format)
    worksheet.write("I{}".format(index + 2), "{0:,.2f}".format(total_length), merge_format)
    worksheet.write("H{}".format(index + 2), "Grand Total", merge_format)
    worksheet.write("I{}".format(index + 2), "{:,.2f}".format(total_length), merge_format)
    worksheet.write("N{}".format(index + 2), "(a) ALUM", None)
    worksheet.write("N{}".format(index + 3), "(b) BUNDLE", None)
    worksheet.write("N{}".format(index + 4), "(i) IDLE", None)
    worksheet.write("N{}".format(index + 5), "Structure Type:SSP - Single Steel Poles; SWP - Single Wood Poles; T - Steel Towers; Other - Multi_Pole Structures or Other Materials", None)
    worksheet.set_column('L:L', 150, cell_format)
    worksheet.set_column('D:D', 35, cell_format)
    worksheet.set_column('N:N', 80, cell_format)
    apply_header_format(workbook, worksheet, df, 0, 0)
    worksheet.write(1, 5, None, None)
    worksheet.write("J0", None, None)


def format_oh_kw_sheet(workbook, worksheet, df):
    """
    This function will format oh_kv Sheet in FERC excel book.
    :param workbook: Workbook Object
    :param worksheet: Oh Conductor Sheet Object for applying sheet level styles
    :param df: Data Frame object
    :return: None

    """
    merge_format = workbook.add_format({
        'bold': 1, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'fg_color': '#efefef'
    })

    cell_format = workbook.add_format()
    cell_format.set_align('center')
    cell_format.set_align('vcenter')
    #cell_format.set_font_name('Times New Roman')

    number_format = workbook.add_format()
    number_format.set_align('center')
    number_format.set_align('vcenter')
    number_format.set_num_format(0x03)

    index = df.shape[0] + 1

    worksheet.set_column('A:C', 30, cell_format)
    worksheet.set_column('C:C', 30, number_format)
    apply_border(workbook, worksheet, "A1:C{}".format(index))

    total_length_mi = df['SUM_Mi'].sum()
    total_length_freq = df['FREQUENCY'].sum()

    worksheet.write("B{}".format(index + 2), total_length_freq, merge_format)
    worksheet.write("C{}".format(index + 2), "{0:,.2f}".format(total_length_mi), merge_format)
    apply_header_format(workbook, worksheet, df, 0, 0)

def format_ug_conductor_sheet(workbook, worksheet, df):
    """
    This function will format ug_conductor Sheet in FERC excel book.
    :param workbook: Workbook Object
    :param worksheet: Ug Conductor Sheet Object for applying sheet level styles
    :param df: Data Frame object
    :return: None

    """
    merge_format = workbook.add_format({
        'bold': 1, 'border': 1,
        'align': 'center', 'valign': 'vcenter',
        'fg_color': '#efefef'
    })
    index = df.shape[0] + 1
    cell_format = workbook.add_format()
    cell_format.set_align('center')
    cell_format.set_align('vcenter')
    #cell_format.set_font_name('Times New Roman')
    worksheet.set_column('A:N', 30, cell_format)
    apply_border(workbook, worksheet, "A1:M{}".format(index))
    total_length = df['CIRCUIT_MI'].sum()
    worksheet.write("H{}".format(index + 2), "Grand Total", merge_format)
    worksheet.write("I{}".format(index + 2), "{0:,.2f}".format(total_length), merge_format)
    worksheet.write("H{}".format(index + 2), "Grand Total", merge_format)
    worksheet.write("I{}".format(index + 2), "{:,.2f}".format(total_length), merge_format)
    worksheet.write("N{}".format(index + 2), "(a) ALUM", None)
    worksheet.write("N{}".format(index + 4), "(i) IDLE", None)
    worksheet.write("N{}".format(index + 5), "Structure Type:N/A for UG ", None)
    worksheet.set_column('L:L', 30, cell_format)
    worksheet.set_column('D:D', 35, cell_format)
    apply_header_format(workbook, worksheet, df, 0, 0)
    worksheet.write(1, 5, None, None)
    worksheet.write("J0", None, None)


def format_ferc_report(worksheet, old_start_row_index, start_row_index, max_rows=0, job_year=2017):
    ft1 = Font(name='Arial', size=10)
    columns = "ABCDEFGHIJKLMNOPRS"
    worksheet.cell(row=old_start_row_index-2, column=1).value = "Summary of Lines"
    worksheet.cell(row=old_start_row_index-1, column=1).value = "listed individually above"
    worksheet.cell(row=old_start_row_index, column=1).value = "Towers & Poles"
    worksheet.cell(row=start_row_index, column=1).value = "Other Undground"
    worksheet.cell(row=start_row_index+1, column=1).value = "Transmission Lines"
    worksheet.cell(row=3, column=5).value = "End of {}/Q4".format(job_year)
    for row_num in (range(22, max_rows)):
        for col in list(columns):
            c = worksheet['{}{}'.format(col, row_num)]
            c.font = ft1
